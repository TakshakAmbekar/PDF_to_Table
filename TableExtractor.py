import re
import pandas as pd
import pdfplumber
from babel.numbers import format_number
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers


#Extract text from PDF file
def extract_text_from_pdf(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text

#Parse the complete PDF text
def parse_bank_statement(text):
    # Split the text into lines
    lines = text.split('\n')
    
    # Initialize list to hold transactions
    transactions = []
    
    # Variable to hold the current transaction being processed
    current_transaction = None
    
    # Regular expression patterns
    date_pattern = r"(?i)\d{2}-[a-z]{3}-\d{4}"  #(?i) flag to make the regular expression case insensitive
    amount_balance_pattern = r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2}[A-Za-z]{2})"  # e.g., "25,000.00 30,38,234.66Dr"
    balance_pattern = r"([\d,]+\.\d{2}[A-Za-z]{2})"  # e.g., "30,63,234.66Dr"
    
    # Process each line
    for line in lines:
        line = line.strip()
        if not line:
            continue  # Skip empty lines
        
        # Check if the line starts with a date
        if re.match(date_pattern, line):
            # If there's a previous transaction, append it
            if current_transaction:
                transactions.append(current_transaction)
            
            # Extract date and description
            date = re.search(date_pattern, line).group(0)
            description = line[len(date):].strip()
            # Initialize new transaction
            current_transaction = {'date': date, 'description': description, 'amount': '', 'balance': ''}
        
        # If we have a current transaction, look for amount and balance
        elif current_transaction:
            # Try to match a transaction line with both amount and balance
            match = re.search(amount_balance_pattern, line)
            if match:
                current_transaction['amount'] = match.group(1)  # e.g., "25,000.00"
                current_transaction['balance'] = match.group(2)  # e.g., "30,38,234.66Dr"
                transactions.append(current_transaction)
                current_transaction = None
            else:
                # Try to match the B/F entry
                match = re.search(balance_pattern, line)
                if match:
                    current_transaction['balance'] = match.group(1)  # e.g., "30,63,234.66Dr"
                    transactions.append(current_transaction)
                    current_transaction = None
                # If no match, append to description (handles multi-line descriptions)
                else:
                    current_transaction['description'] += ' ' + line
    
    # Append the last transaction if it exists
    if current_transaction:
        transactions.append(current_transaction)
    
    # Create a DataFrame from the transactions
    df = pd.DataFrame(transactions, columns=['date', 'description', 'amount', 'balance'])
    return df

if __name__ == "__main__":
    # Path to the pdf file
    pdf_path = "test1.pdf"
    excel_path = "statement.xlsx"
    
    # Parse the statement
    df = parse_bank_statement(extract_text_from_pdf(pdf_path))

    # Convert 'date' column to datetime format
    df['date'] = pd.to_datetime(df['date'], format='%d-%b-%Y').dt.strftime('%d-%b-%Y')

    # Remove commas and extract numeric values from 'amount' and 'balance'
    df['amount'] = df['amount'].str.replace(',', '').str.extract(r'([\d.]+)')[0]
    df['balance'] = df['balance'].str.replace(',', '').str.extract(r'([\d.]+)')[0]

    # Convert to float after cleaning
    df['amount'] = df['amount'].astype(float).round(2)
    df['balance'] = df['balance'].astype(float).round(2)

    # # Format as Indian locale string
    # df['amount'] = df['amount'].apply(lambda x: format_number(x, locale='en_IN') if pd.notnull(x) else "")
    # df['balance'] = df['balance'].apply(lambda x: format_number(x, locale='en_IN') if pd.notnull(x) else "")

    # Save to Excel with formatting using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Write DataFrame to Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Apply number formatting to 'amount' and 'balance' columns
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):  # Columns C and D
        for cell in row:
            cell.number_format = numbers.FORMAT_NUMBER_00  # Ensures 2 decimal places

    # Define table range dynamically
    table_ref = f"A1:D{ws.max_row}"

    # Create a table
    table = Table(displayName="BankTransactions", ref=table_ref)

    # Apply a table style
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style

    ws.add_table(table)

    # Save workbook
    wb.save(excel_path)

    print(f"Excel file '{excel_path}' created successfully.")