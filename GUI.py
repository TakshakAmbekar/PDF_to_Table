import re
import os
import pandas as pd
import pdfplumber
from babel.numbers import format_number
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

class PDFTableExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF Table Extractor")
        self.root.geometry("600x300")
        self.root.minsize(600, 300)
        
        # Variables
        self.pdf_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        self.date_format = tk.StringVar(value="%d-%b-%Y")
        self.format_example = tk.StringVar(value="Example: 01-Jan-2023")
        
        # Default headers (hidden from user)
        self.default_headers = ["Date", "Description", "Amount", "Balance"]
        
        # Default regex patterns (hidden from user)
        self.date_pattern = r"(?i)\d{2}-[a-z]{3}-\d{4}"
        self.amount_balance_pattern = r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2}[A-Za-z]{2})"
        self.balance_pattern = r"([\d,]+\.\d{2}[A-Za-z]{2})"
        
        # Create main frame
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Application title
        title_label = ttk.Label(main_frame, text="PDF Bank Statement Table Extractor", font=("Helvetica", 14, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Create file selection frame
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill=tk.X, pady=5)
        
        # PDF file selection
        ttk.Label(file_frame, text="PDF File:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(file_frame, textvariable=self.pdf_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="Browse...", command=self.browse_pdf).grid(row=0, column=2, padx=5, pady=5)
        
        # Excel file selection
        ttk.Label(file_frame, text="Excel Output:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(file_frame, textvariable=self.excel_path, width=50).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="Browse...", command=self.browse_excel).grid(row=1, column=2, padx=5, pady=5)
        
        # Date format selection
        date_frame = ttk.Frame(main_frame)
        date_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(date_frame, text="Date Format:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        
        # More date formats
        date_formats = [
            "%d-%b-%Y",    # 01-Jan-2023
            "%d/%m/%Y",    # 01/01/2023
            "%m/%d/%Y",    # 01/01/2023
            "%Y-%m-%d",    # 2023-01-01
            "%d-%m-%Y",    # 01-01-2023
            "%d.%m.%Y",    # 01.01.2023
            "%b %d, %Y",   # Jan 01, 2023
            "%d %b %Y",    # 01 Jan 2023
            "%Y/%m/%d"     # 2023/01/01
        ]
        
        date_combo = ttk.Combobox(date_frame, textvariable=self.date_format, values=date_formats, width=15)
        date_combo.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        date_combo.bind("<<ComboboxSelected>>", self.update_date_example)
        
        ttk.Label(date_frame, textvariable=self.format_example).grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        
        # Run button
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=20)
        
        run_button = ttk.Button(button_frame, text="Extract Table and Generate Excel", command=self.run_extraction, width=30)
        run_button.pack(pady=10)
        
        # Initialize date example
        self.update_date_example(None)
    
    def browse_pdf(self):
        filename = filedialog.askopenfilename(
            title="Select PDF File",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if filename:
            self.pdf_path.set(filename)
            # Auto-set Excel path
            base_name = os.path.splitext(os.path.basename(filename))[0]
            excel_dir = os.path.dirname(filename)
            self.excel_path.set(os.path.join(excel_dir, f"{base_name}_extracted.xlsx"))
    
    def browse_excel(self):
        filename = filedialog.asksaveasfilename(
            title="Save Excel File As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if filename:
            self.excel_path.set(filename)
    
    def update_date_example(self, event):
        """Update the example date format based on the selected format"""
        date_format = self.date_format.get()
        
        # Sample date to format (January 1, 2023)
        from datetime import datetime
        sample_date = datetime(2023, 1, 1)
        
        try:
            example = sample_date.strftime(date_format)
            self.format_example.set(f"Example: {example}")
        except:
            self.format_example.set("Invalid format")
    
    def run_extraction(self):
        """Run the extraction process"""
        if not self.pdf_path.get():
            messagebox.showerror("Error", "Please select a PDF file.")
            return
        
        if not self.excel_path.get():
            messagebox.showerror("Error", "Please specify an Excel output file.")
            return
        
        try:
            # Extract text from PDF
            text = self.extract_text_from_pdf(self.pdf_path.get())
            
            # Parse the text
            df = self.parse_bank_statement(text)
            
            if df.empty:
                messagebox.showwarning("Warning", "No transactions found. Please check the PDF format.")
                return
            
            # Convert date format
            input_format = self.get_input_date_format()
            output_format = self.date_format.get()
            
            try:
                df['Date'] = pd.to_datetime(df['Date'], format=input_format).dt.strftime(output_format)
            except Exception as e:
                messagebox.showwarning("Warning", f"Date conversion failed: {str(e)}\nProceeding with original dates.")
            
            # Clean amount and balance columns
            df['Amount'] = df['Amount'].str.replace(',', '').str.extract(r'([\d.]+)')[0]
            df['Balance'] = df['Balance'].str.replace(',', '').str.extract(r'([\d.]+)')[0]
            
            # Convert to float
            df['Amount'] = df['Amount'].astype(float).round(2)
            df['Balance'] = df['Balance'].astype(float).round(2)
            
            # Save to Excel
            self.save_to_excel(df, self.excel_path.get())
            
            messagebox.showinfo("Success", f"Excel file created successfully at:\n{self.excel_path.get()}")
            
            # Ask if user wants to extract another file
            self.prompt_continue()
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
    
    def prompt_continue(self):
        """Ask if the user wants to extract another file or finish"""
        response = messagebox.askyesno(
            "Extraction Complete", 
            "Do you want to extract a table from another PDF file?",
            icon=messagebox.QUESTION
        )
        
        if response:
            # Reset the form for a new extraction
            self.pdf_path.set("")
            self.excel_path.set("")
        else:
            # Close the application
            self.root.destroy()
    
    def extract_text_from_pdf(self, pdf_path):
        """Extract text from PDF file"""
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() + "\n"
        return text
    
    def parse_bank_statement(self, text):
        """Parse the bank statement text to extract transactions"""
        # Split the text into lines
        lines = text.split('\n')
        
        # Initialize list to hold transactions
        transactions = []
        
        # Variable to hold the current transaction being processed
        current_transaction = None
        
        # Process each line
        for line in lines:
            line = line.strip()
            if not line:
                continue  # Skip empty lines
            
            # Check if the line starts with a date
            if re.match(self.date_pattern, line):
                # If there's a previous transaction, append it
                if current_transaction:
                    transactions.append(current_transaction)
                
                # Extract date and description
                date = re.search(self.date_pattern, line).group(0)
                description = line[len(date):].strip()
                # Initialize new transaction
                current_transaction = {'Date': date, 'Description': description, 'Amount': '', 'Balance': ''}
            
            # If we have a current transaction, look for amount and balance
            elif current_transaction:
                # Try to match a transaction line with both amount and balance
                match = re.search(self.amount_balance_pattern, line)
                if match:
                    current_transaction['Amount'] = match.group(1)  # e.g., "25,000.00"
                    current_transaction['Balance'] = match.group(2)  # e.g., "30,38,234.66Dr"
                    transactions.append(current_transaction)
                    current_transaction = None
                else:
                    # Try to match the B/F entry
                    match = re.search(self.balance_pattern, line)
                    if match:
                        current_transaction['Balance'] = match.group(1)  # e.g., "30,63,234.66Dr"
                        transactions.append(current_transaction)
                        current_transaction = None
                    # If no match, append to description (handles multi-line descriptions)
                    else:
                        current_transaction['Description'] += ' ' + line
        
        # Append the last transaction if it exists
        if current_transaction:
            transactions.append(current_transaction)
        
        # Create a DataFrame from the transactions
        df = pd.DataFrame(transactions, columns=self.default_headers)
        return df
    
    def get_input_date_format(self):
        """Determine the input date format based on the regex pattern"""
        # Default to the format in the original script
        return "%d-%b-%Y"
    
    def save_to_excel(self, df, excel_path):
        """Save the DataFrame to an Excel file with formatting"""
        # Save to Excel with formatting using openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Write DataFrame to Excel
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply number formatting to 'amount' and 'balance' columns
        amount_col = 3  # Assuming amount is the 3rd column (C)
        balance_col = 4  # Assuming balance is the 4th column (D)
        
        # Apply formatting
        for row in ws.iter_rows(min_row=2, min_col=amount_col, max_col=amount_col):
            for cell in row:
                cell.number_format = numbers.FORMAT_NUMBER_00
        
        for row in ws.iter_rows(min_row=2, min_col=balance_col, max_col=balance_col):
            for cell in row:
                cell.number_format = numbers.FORMAT_NUMBER_00
        
        # Define table range dynamically
        table_ref = f"A1:{chr(64 + len(df.columns))}{ws.max_row}"
        
        # Create a table
        table = Table(displayName="BankTransactions", ref=table_ref)
        
        # Apply a table style
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        ws.add_table(table)
        
        # Save workbook
        wb.save(excel_path)

if __name__ == "__main__":
    root = tk.Tk()
    app = PDFTableExtractorGUI(root)
    root.mainloop()

print("This code creates a simplified GUI for extracting tabular data from PDF files.")
print("Run this script to launch the application.")